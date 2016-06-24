import numpy as np
import matplotlib.pyplot as plt

N = 2 # number of observations, with j
S = 2 # number of stars, with i

W = [None for i in range(S)]
#W[0] is star 0 and has the format [xarr, yarr, oarr, earr, m, delm]
x0 = [0,1,3,4,2,4,0,1]
y0 = [3,4,0,1,3,4,0,2]
o0 = [25.5,25,24,24,25,25,25,70]
e0 = np.array([.12,.15,.1,.1,.12,.12])

def magest(o,e,N):
    # o and e are arrays of unbiased estimate of unknown magnitude
    # and e is error
    o = np.asarray(o)
    e = np.asarray(e)
    top = np.sum(o/e**2)
    bot = np.sum(1/e**2)
    m = top/bot # optimal estimate of mag of star i (eq2)
    delm = bot**(-.5) # formal random error on m of star i (eq3)
    return m, delm
#m0,delm0 = magest(o0,e0,N)
#W[0] = [x0,y0,o0,e0,m0,delm0]
#print W

def sigmaclip(a, low = 3, high = 3):
    # copied exactly from scipy.stats.sigmaclip with some variation to keep 
    # account for the index(es) that are being removed
    c = np.asarray(a).ravel()
    remove_arr = np.array([]) #indicies that have been removed
    delta = 1
    while delta:
        c_std = c.std()
        c_mean = c.mean()
        size = c.size
        critlower = c_mean - c_std*low
        critupper = c_mean + c_std*high
        removetemp = np.where(c < critlower) and np.where(c > critupper)
        remove_arr = np.append(remove_arr, removetemp)
        c = np.delete(c, removetemp)
        delta = size-c.size
    return c, remove_arr
    
def func(x,y):
    #the function we want to optimize
    return [x**2, y**2,x*y,x,y,1+x*0]
    
def getfit(x,y,z, pixelx, pixely):
    """
    This is just for one star/object!
    Parameters:
    x:      array of x pixel values
    y:      array of y pixel values
    z:      array of magnitude values at a certian pixel (x,y)
    pixelx: array of all ranging x pixel values 
    pixely: array of all ranging y pixel values
    
    Returns:
    zfit:   the fitted magnitude values at the pixels given by the x and y input 
            arrays
    Zfit:   the fitted magnitude values at all the pixels of values given by 
            pixelx and pixely (will be used to create a 3D plot)
    X:      the mesh x array (will be used to create a 3D plot)
    Y:      the mesy y array (will be used to create a 3D plot)
    rsum:   the sum of residuals (LLS works by minimizing this value) :: if the 
            value is an empty array, the fit is perfectly matched to data
    resarr: an array of the fitted z values compared to the input z values 
            (the absolute value of the error)
    """
    
    x = np.asarray(x)
    y = np.asarray(y)
    z = np.asarray(z)
    
    # sigma clipping: to remove anything that is above/below a certain sigma
    z, remove_arr = sigmaclip(z, low=2.5, high=2.5)
    x = np.delete(x, remove_arr)
    y = np.delete(y, remove_arr)
    
    X,Y = np.meshgrid(pixelx, pixely, sparse = True, copy=False) #why copy false??
    f = func(x,y)
    fmesh = func(X,Y)
    A = np.array(f).T
    B = z
    coeff, rsum, rank, s = np.linalg.lstsq(A, B)
    
    zfit = np.zeros(len(x))
    Zfit = [[0 for i in pixelx] for j in pixely]
    k = 0
    while k < len(coeff):
        zfit += coeff[k]*f[k]
        Zfit += coeff[k]*fmesh[k]
        k+=1
    #zfit = coeff[0]*x**2 + coeff[1]*y**2 + ... === coeff[0]*f[0] + ...
    #Zfit = coeff[0]*X**2 + coeff[1]*Y**2 + ... === coeff[0]*fmesh[0] + ...
    
    def get_res(z, zfit):
        # returns an array of the error in the fit
        return np.abs(zfit-z)
    resarr = get_res(z,zfit)
    return [zfit, Zfit, x, y, X, Y, rsum, resarr, remove_arr]
    
zfit, Zfit, x, y, X, Y, rsum, resarr, remove_arr = getfit(x0,y0,o0, range(5), range(5))
print "Residual sum: ", rsum
print "Number of removed points: " , len(remove_arr)
print "Ratio of removed points to original total points: " , np.double(len(remove_arr))/len(x0)

def plot3d(x,y,z,X,Y,Z):    
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    ax.plot_wireframe(X,Y,Z, rstride=1, cstride=1)
    ax.scatter(x,y,z, s= 50)
    ax.set_xlabel('X Pixel')
    ax.set_ylabel('Y Pixel')
    return fig

figtest = plot3d(x,y,np.delete(o0,remove_arr),X,Y,Zfit)
plt.show(figtest)

CHIP1YLEN = 2048
CHIP2YLEN = 2048
filename = '/Users/dkossakowski/Desktop/testdata.xlsx'
def get_params(filename, chipstring = 'C', xstring = 'D', ystring = 'E', mstring = 'F'):
    import openpyxl
    from openpyxl.cell import column_index_from_string
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')
    
    chip_arr = np.array([])
    x_arr = np.array([])
    y_arr = np.array([])
    m_arr = np.array([])
    chipcol = column_index_from_string(chipstring)
    xcol = column_index_from_string(xstring)
    ycol = column_index_from_string(ystring)
    mcol = column_index_from_string(mstring)
    for i in np.arange(1,10):
        chip = sheet.cell(row=i, column=chipcol).value
        chip_arr = np.append(chip_arr, chip)
        xval = sheet.cell(row=i, column=xcol).value
        x_arr = np.append(x_arr, xval)
        yval = sheet.cell(row=i, column=ycol).value
        if chip == 1. or chip == 1:
            yval += CHIP2YLEN
        y_arr = np.append(y_arr, yval)
        mval = sheet.cell(row=i, column=mcol).value
        m_arr = np.append(m_arr, mval)
    return [x_arr, y_arr, m_arr]
    
x, y, m = get_params(filename)
#zfit, Zfit, x, y, X, Y, rsum, res_arr, remove_arr = getfit(x,y,m, np.arange(0,4096,256), range(0,4096,256))
#print "residual sum: ", rsum
#figtest1 = plot3d(x,y,np.delete(m,remove_arr),X,Y,Zfit)   
#plt.show(figtest1)






    
    